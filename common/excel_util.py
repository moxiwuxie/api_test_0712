#!/usr/bin/python
# -*- coding: UTF-8 -*-
import openpyxl
import os
def write_excel(table_name,sheet_name,coordinate_name):
  wb = openpyxl.load_workbook(table_name)
  ws = wb[sheet_name]
  return ws[coordinate_name].value
def save_excel(table_name,sheet_name,coordinate_name,data):
  wb = openpyxl.load_workbook(table_name)
  ws = wb[sheet_name]
  ws[coordinate_name] = data
  wb.save(table_name)
  print(coordinate_name + " " + str(data) + "save success")

